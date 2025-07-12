import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from dotenv import load_dotenv
import yagmail

# Carrega as variáveis do .env
load_dotenv()
email_remetente = os.getenv("EMAIL")
senha = os.getenv("SENHAEMAIL")

# Conecta ao servidor de e-mail
yag = yagmail.SMTP(user=email_remetente, password=senha)

# Carrega a planilha de resultados
wb = load_workbook("resultado_final.xlsx")
sheet = wb.active # pega somente as abas ativas da planilha.

# Lê os dados da planilha (a partir da 2ª linha)
for row in sheet.iter_rows(min_row=2, values_only=True):
    nome, media, situacao, email_destino = row # cria 4 variaveis com os 4 valores da linha.

    # Cria um PDF novo (individual)
    nome_arquivo = f"relatorio_{nome}.pdf" # f permite criar variaveis dentro da String.
    pdf = canvas.Canvas(nome_arquivo, pagesize=A4)
    largura, altura = A4 # vai receber largura ≈ 595, altura ≈ 842.

    # --- ADICIONANDO UMA IMAGEM ---
    logo = ImageReader("logo.png")

    # Local da imagem
    # x = centro - (largura da imagem / 2)
    largura_logo = 6 * cm
    altura_logo = 2 * cm
    x_logo = (largura - largura_logo) / 2  # centraliza
    y_logo = altura - altura_logo - 20  # 20 pontos de margem do topo

    pdf.drawImage(logo, x_logo, y_logo, width=largura_logo, height=altura_logo)

    # Título
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(200, altura - 100, "Relatório de Notas")

    # Cabeçalho
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, altura - 200, "Nome")
    pdf.drawString(250, altura - 200, "Média")
    pdf.drawString(350, altura - 200, "Situação")

    # Dados do aluno
    y = altura - 220 # y = 722

    # Ajusta a fonte e cor se reprovado
    if situacao == "Reprovado":
        pdf.setFillColorRGB(1, 0, 0)  # vermelho
    else:
        pdf.setFillColorRGB(0, 0, 0)  # preto

    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, y, str(nome))
    pdf.drawString(250, y, str(media))
    pdf.drawString(350, y, situacao)

    # Finaliza o PDF
    pdf.save()

    # Envia o e-mail com o PDF individual
    assunto = f"Relatório Escolar de {nome}"
    corpo = f"Olá {nome},\n\nSegue em anexo seu relatório escolar com o resultado final.\n\nAtenciosamente,\nCoordenação"

    yag.send(
        to=email_destino,
        subject=assunto,
        contents=corpo,
        attachments=nome_arquivo
    )

    print(f"✅ E-mail enviado para {nome} ({email_destino}) com anexo: {nome_arquivo}")
